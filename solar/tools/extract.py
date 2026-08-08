"""売電結果2026.xlsx の「累計」シートから、アプリが使う js/data.js を作る。

    python3 tools/extract.py 売電結果2026.xlsx > js/data.js

シートの並び（行＝月、列＝年）はこのファイルの PLANTS と salescol が持っている。
シートの形が変わったらここだけ直せばよい。

必要なもの: pip install openpyxl
"""
import json, sys, openpyxl
from openpyxl.utils import column_index_from_string as ci

src = sys.argv[1] if len(sys.argv) > 1 else '売電結果2026.xlsx'
wb = openpyxl.load_workbook(src, data_only=True)
ws = wb['累計']

YEARS = list(range(2018, 2027))          # 2018..2026
def salescol(y): return ci('E') + (y - 2018) * 2   # E,G,I,K,M,O,Q,S,U

PLANTS = [
    dict(id='ichihara', name='市原発電所', pref='千葉', kw=76, unit=21,
         start='2018-03', bank='ジャックス', rate=2.15, price=17500000,
         row0=6, profitRow=19, simCol='C',
         costCols=[('ローン', 'W'), ('PC電気代', 'X'), ('草刈り', 'Y'), ('償却資産税', 'Z')]),
    dict(id='futtsu', name='富津発電所', pref='千葉', kw=99, unit=14,
         start='2021-01', bank='日本政策金融公庫', rate=2.25, price=17000000,
         row0=24, profitRow=37, simCol=None,
         costCols=[('ローン', 'W'), ('草刈り・電気', 'X'), ('土地代', 'Y'), ('償却資産税', 'Z')]),
    dict(id='takehara', name='竹原発電所', pref='広島', kw=97, unit=18,
         start='2018-12', bank='ジャックス', rate=2.2, price=16700000,
         row0=42, profitRow=55, simCol='C',
         costCols=[('ローン', 'W'), ('草刈り', 'X'), ('電気', 'Y'), ('償却資産税', 'Z')]),
]

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)):
        return int(round(v))
    return None

out = {'source': '売電結果2026.xlsx（累計シート）', 'updated': '2026-07-24', 'years': YEARS, 'plants': []}

for p in PLANTS:
    r0 = p['row0']
    years = {}
    for y in YEARS:
        c = salescol(y)
        sales = [num(ws.cell(r0 + m, c).value) for m in range(12)]
        kwh   = [num(ws.cell(r0 + m, c + 1).value) for m in range(12)]
        if any(v for v in sales) or any(v for v in kwh):
            years[str(y)] = {'sales': sales, 'kwh': kwh}
    sim = None
    if p['simCol']:
        c = ci(p['simCol'])
        sim = {'sales': [num(ws.cell(r0 + m, c).value) for m in range(12)],
               'kwh':   [num(ws.cell(r0 + m, c + 1).value) for m in range(12)]}
    # 月ごとの費用
    costs = []
    for m in range(12):
        items = []
        for label, col in p['costCols']:
            v = num(ws.cell(r0 + m, ci(col)).value)
            if v: items.append({'name': label, 'amount': v})
        costs.append(items)
    # 年ごとの確定利益
    conf = {}
    for y in YEARS:
        v = num(ws.cell(p['profitRow'], salescol(y)).value)
        if v is not None: conf[str(y)] = v
    q = dict(p); q.pop('row0'); q.pop('profitRow'); q.pop('simCol'); q.pop('costCols')
    q['costLabels'] = [l for l, _ in p['costCols']]
    q['years'] = years
    q['sim'] = sim
    q['costs'] = costs
    q['confirmed'] = conf
    out['plants'].append(q)

HEADER = '''/* data.js \u2014 \u58f2\u96fb\u7d50\u679c2026.xlsx \u306e\u300c\u7d2f\u8a08\u300d\u30b7\u30fc\u30c8\u304b\u3089\u53d6\u308a\u51fa\u3057\u305f\u5b9f\u7e3e\u3002
 *
 *   \u58f2\u96fb\u984d\uff08\u5186\uff09\u3068\u767a\u96fb\u91cf\uff08kWh\uff09\u3092\u3001\u767a\u96fb\u6240\u3054\u3068\u30fb\u5e74\u3054\u3068\u306b 1\u6708\u301c12\u6708\u306e12\u500b\u4e26\u3073\u3067\u6301\u3064\u3002
 *   \u672a\u8a18\u5165\u306e\u6708\u306f null\u3002\u6708\u3054\u3068\u306e\u8cbb\u7528\uff08\u30ed\u30fc\u30f3\u30fb\u8349\u5208\u308a\u30fb\u96fb\u6c17\u30fb\u511f\u5374\u8cc7\u7523\u7a0e\u306a\u3069\uff09\u3082
 *   \u540c\u3058\u30b7\u30fc\u30c8\u306e W\u301cZ \u5217\u304b\u3089\u53d6\u308a\u8fbc\u3093\u3067\u3042\u308b\u3002
 *
 *   \u3053\u3053\u306f\u300c\u5143\u5e33\u300d\u3002\u30ab\u30e1\u30e9\u3084\u624b\u5165\u529b\u3067\u8db3\u3057\u305f\u5206\u306f store.js \u5074\u306b\u5225\u3067\u8caf\u3081\u3001
 *   \u8868\u793a\u3059\u308b\u3068\u304d\u306b\u91cd\u306d\u308b\u3002\u5143\u306e\u30b7\u30fc\u30c8\u3092\u53d6\u308a\u8fbc\u307f\u76f4\u3057\u3066\u3082\u3001\u8ffd\u8a18\u304c\u6d88\u3048\u306a\u3044\u3088\u3046\u306b\u3059\u308b\u305f\u3081\u3002
 *
 *   \u203b \u3053\u306e\u30d5\u30a1\u30a4\u30eb\u306f tools/extract.py \u304c xlsx \u304b\u3089\u4f5c\u308b\u3002\u624b\u3067\u76f4\u3055\u306a\u3044\u3053\u3068\u3002
 */
window.SOLAR_BASE = '''

sys.stdout.write(HEADER + json.dumps(out, ensure_ascii=False, separators=(',', ':')) + ';\n')

# \u691c\u7b97\uff1a3\u57fa\u5408\u8a08\u3068 \u7d2f\u8a08\u30b7\u30fc\u30c8\u306e\u5408\u8a08\u884c\u3092\u7a81\u304d\u5408\u308f\u305b\u3066\u3001\u5408\u308f\u306a\u3051\u308c\u3070\u6b62\u3081\u308b
for y, ref in ((2024, 'Q71'), (2025, 'S71'), (2026, 'U71')):
    tot = 0
    for p in out['plants']:
        d = p['years'].get(str(y))
        if d:
            tot += sum(v or 0 for v in d['sales'])
    want = ws[ref].value
    if want is not None and round(want) != tot:
        sys.exit('%d\u5e74\u306e3\u57fa\u5408\u8a08\u304c\u30b7\u30fc\u30c8\u306e%s\u3068\u5408\u3044\u307e\u305b\u3093: %d != %s' % (y, ref, tot, want))
    print('%d\u5e74 3\u57fa\u5408\u8a08 %s \u2713 (%s)' % (y, format(tot, ','), ref), file=sys.stderr)
